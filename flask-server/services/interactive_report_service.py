import base64
import io
import math
import os
import re
from typing import Any, Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import nibabel as nib
import numpy as np
import pandas as pd
import SimpleITK as sitk
from openpyxl import load_workbook

from constants import Constants


GLOSSARY: Dict[str, Dict[str, str]] = {
    "pancreatic duct dilation": {
        "label": "Pancreatic duct dilation",
        "definition": "Widening of the pancreatic duct, which can be associated with obstruction, inflammation, or a pancreatic mass.",
        "example": "In pancreatic cancer workflows, duct dilation is often reviewed together with lesion location and upstream gland changes."
    },
    "hypoattenuating lesion": {
        "label": "Hypoattenuating lesion",
        "definition": "A region that appears darker than surrounding tissue on CT, meaning it attenuates fewer X-rays.",
        "example": "Pancreatic adenocarcinoma is often described as hypoattenuating relative to the surrounding pancreas on contrast-enhanced CT."
    },
    "malignant mass": {
        "label": "Malignant mass",
        "definition": "A suspicious tissue abnormality concerning for cancer, based on imaging appearance, clinical context, or pathology.",
        "example": "In a research viewer, this term should be interpreted as descriptive or suggestive unless pathology confirmation is available."
    },
    "pancreatic lesion": {
        "label": "Pancreatic lesion",
        "definition": "A focal abnormality involving pancreatic tissue. It may represent benign, inflammatory, cystic, or malignant pathology.",
        "example": "Interactive reports can link a pancreatic lesion sentence directly to the corresponding axial slice and segmentation contour."
    },
    "head of the pancreas": {
        "label": "Head of the pancreas",
        "definition": "The broad rightward portion of the pancreas adjacent to the duodenum.",
        "example": "Lesions in the pancreatic head are often clinically important because they may affect the common bile duct or pancreatic duct."
    },
    "body of the pancreas": {
        "label": "Body of the pancreas",
        "definition": "The central portion of the pancreas between the head and tail.",
        "example": "Body lesions may be more subtle and sometimes present later than pancreatic head lesions."
    },
    "tail of the pancreas": {
        "label": "Tail of the pancreas",
        "definition": "The leftward tapering end of the pancreas extending toward the spleen.",
        "example": "Tail lesions may be missed if attention is focused only on the pancreatic head."
    },
}

LABEL_NAME_TO_VALUE = {name: idx for idx, name in Constants.PREDEFINED_LABELS.items()}


def get_pants_case_id(case_id: str) -> str:
    case_id = str(case_id)
    return "PanTS_" + case_id.zfill(8)


def _training_subfolder(case_id: str) -> Tuple[str, str]:
    cid = int(case_id)
    image_sub = "ImageTr" if cid < 9000 else "ImageTe"
    label_sub = "LabelTr" if cid < 9000 else "LabelTe"
    return image_sub, label_sub


def get_case_paths(case_id: str) -> Dict[str, str]:
    image_sub, label_sub = _training_subfolder(case_id)
    pants_id = get_pants_case_id(case_id)
    ct_path = os.path.join(
        Constants.PANTS_PATH, "data", image_sub, pants_id, Constants.MAIN_NIFTI_FILENAME
    )
    combined_labels_path = os.path.join(
        Constants.PANTS_PATH, "data", label_sub, pants_id, Constants.COMBINED_LABELS_NIFTI_FILENAME
    )
    return {
        "ct_path": ct_path,
        "combined_labels_path": combined_labels_path,
        "pants_id": pants_id,
    }


def load_case_metadata(case_id: str) -> Dict[str, Any]:
    """
    Loads the best available case metadata.
    Priority:
    1) sessions/<id>/info.csv if present
    2) PanTS/data/metadata.xlsx if present
    """
    result: Dict[str, Any] = {
        "age": None,
        "sex": None,
        "narrative_report": "",
        "structured_report": "",
        "scanner_model": None,
        "study_date": None,
    }

    session_csv = os.path.join(Constants.SESSIONS_DIR_NAME, str(case_id), "info.csv")
    if os.path.exists(session_csv):
        try:
            df = pd.read_csv(session_csv)
            if len(df) > 0:
                row = df.iloc[0]
                lower_map = {str(col).strip().lower(): col for col in df.columns}

                def _pick(*names: str):
                    for n in names:
                        if n.lower() in lower_map:
                            return row[lower_map[n.lower()]]
                    return None

                result["age"] = _pick("age")
                result["sex"] = _pick("sex")
                result["narrative_report"] = _pick("narrative_report", "narrative report", "report") or ""
                result["structured_report"] = _pick("structured_report", "structured report") or ""
                result["scanner_model"] = _pick("manufacturer_model_name", "scanner_model")
                result["study_date"] = _pick("study_date", "scan_date")
                return result
        except Exception:
            pass

    metadata_xlsx = os.path.join(Constants.PANTS_PATH, "data", "metadata.xlsx")
    if os.path.exists(metadata_xlsx):
        try:
            wb = load_workbook(metadata_xlsx, read_only=True)
            if "PanTS_metadata" in wb.sheetnames:
                sheet = wb["PanTS_metadata"]
                target_id = get_pants_case_id(case_id)
                for row in sheet.iter_rows(values_only=True):
                    if row and row[0] == target_id:
                        result["sex"] = row[4] if len(row) > 4 else None
                        result["age"] = row[5] if len(row) > 5 else None
                        break
        except Exception:
            pass

    return result


def _read_ras_arrays(ct_path: str, combined_labels_path: str) -> Tuple[np.ndarray, np.ndarray, Tuple[float, float, float]]:
    ct_img = sitk.ReadImage(ct_path)
    ct_img = sitk.DICOMOrient(ct_img, "RAS")
    ct_arr = sitk.GetArrayFromImage(ct_img)

    label_img = sitk.ReadImage(combined_labels_path)
    label_img = sitk.DICOMOrient(label_img, "RAS")
    label_arr = sitk.GetArrayFromImage(label_img)

    spacing_xyz = ct_img.GetSpacing()  # x, y, z
    spacing_zyx = (float(spacing_xyz[2]), float(spacing_xyz[1]), float(spacing_xyz[0]))
    return ct_arr, label_arr, spacing_zyx


def _window_ct(slice_2d: np.ndarray, min_hu: int = -150, max_hu: int = 250) -> np.ndarray:
    clipped = np.clip(slice_2d, min_hu, max_hu)
    norm = (clipped - min_hu) / float(max_hu - min_hu)
    return (norm * 255).astype(np.uint8)


def _largest_slice_index(mask_3d: np.ndarray) -> int:
    slice_sums = np.sum(mask_3d > 0, axis=(1, 2))
    return int(np.argmax(slice_sums))


def _bbox_from_mask_slice(mask_slice: np.ndarray, padding: int = 20) -> Optional[Dict[str, int]]:
    coords = np.argwhere(mask_slice > 0)
    if coords.size == 0:
        return None

    min_row, min_col = coords.min(axis=0)
    max_row, max_col = coords.max(axis=0)

    min_row = max(int(min_row) - padding, 0)
    min_col = max(int(min_col) - padding, 0)
    max_row = min(int(max_row) + padding, mask_slice.shape[0] - 1)
    max_col = min(int(max_col) + padding, mask_slice.shape[1] - 1)

    return {
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
    }


def _fig_to_data_url() -> str:
    buffer = io.BytesIO()
    plt.savefig(buffer, format="png", bbox_inches="tight", pad_inches=0)
    plt.close()
    buffer.seek(0)
    encoded = base64.b64encode(buffer.read()).decode("utf-8")
    return f"data:image/png;base64,{encoded}"


def _render_overlay_and_zoom(
    ct_arr: np.ndarray,
    mask_arr: np.ndarray,
    color: str = "red"
) -> Dict[str, Any]:
    idx = _largest_slice_index(mask_arr)
    ct_slice = ct_arr[idx]
    mask_slice = (mask_arr[idx] > 0).astype(np.uint8)

    bbox = _bbox_from_mask_slice(mask_slice)
    if bbox is None:
        raise ValueError("No positive voxels found in mask slice.")

    # Full slice overlay
    display_ct = np.fliplr(_window_ct(ct_slice))
    display_mask = np.fliplr(mask_slice)

    plt.figure(figsize=(6, 6))
    plt.imshow(display_ct, cmap="gray", origin="lower")
    plt.contour(display_mask, colors=color, linewidths=1.5)
    plt.axis("off")
    overlay_url = _fig_to_data_url()

    # Zoomed crop
    zoom_ct = ct_slice[
        bbox["min_row"]:bbox["max_row"] + 1,
        bbox["min_col"]:bbox["max_col"] + 1
    ]
    zoom_mask = mask_slice[
        bbox["min_row"]:bbox["max_row"] + 1,
        bbox["min_col"]:bbox["max_col"] + 1
    ]

    display_zoom_ct = np.fliplr(_window_ct(zoom_ct))
    display_zoom_mask = np.fliplr(zoom_mask)

    plt.figure(figsize=(5, 5))
    plt.imshow(display_zoom_ct, cmap="gray", origin="lower")
    plt.contour(display_zoom_mask, colors=color, linewidths=1.5)
    plt.axis("off")
    zoom_url = _fig_to_data_url()

    return {
        "slice_index": idx,
        "bbox": bbox,
        "overlay_image": overlay_url,
        "zoom_image": zoom_url,
    }


def _extract_binary_mask(label_arr: np.ndarray, label_name: str) -> np.ndarray:
    label_value = LABEL_NAME_TO_VALUE.get(label_name)
    if label_value is None:
        return np.zeros_like(label_arr, dtype=np.uint8)
    return (label_arr == label_value).astype(np.uint8)


def _centroid_zyx(mask_arr: np.ndarray) -> Optional[Tuple[float, float, float]]:
    coords = np.argwhere(mask_arr > 0)
    if coords.size == 0:
        return None
    centroid = coords.mean(axis=0)
    return float(centroid[0]), float(centroid[1]), float(centroid[2])


def _infer_pancreatic_region(label_arr: np.ndarray, lesion_mask: np.ndarray) -> str:
    if lesion_mask.sum() == 0:
        return "pancreas"

    region_candidates = {
        "head of the pancreas": "pancreas_head",
        "body of the pancreas": "pancreas_body",
        "tail of the pancreas": "pancreas_tail",
        "pancreas": "pancreas",
    }

    overlaps = {}
    for region_label, internal_name in region_candidates.items():
        region_mask = _extract_binary_mask(label_arr, internal_name)
        overlaps[region_label] = int(np.sum((lesion_mask > 0) & (region_mask > 0)))

    best_region = max(overlaps, key=overlaps.get)
    return best_region if overlaps[best_region] > 0 else "pancreas"


def _approx_long_axis_mm(mask_arr: np.ndarray, spacing_zyx: Tuple[float, float, float]) -> float:
    coords = np.argwhere(mask_arr > 0)
    if coords.size == 0:
        return 0.0

    z_min, y_min, x_min = coords.min(axis=0)
    z_max, y_max, x_max = coords.max(axis=0)

    dz = (z_max - z_min + 1) * spacing_zyx[0]
    dy = (y_max - y_min + 1) * spacing_zyx[1]
    dx = (x_max - x_min + 1) * spacing_zyx[2]
    return round(max(dx, dy, dz), 1)


def _term_hits(text: str) -> List[str]:
    lowered = text.lower()
    hits = []
    for term in GLOSSARY.keys():
        if term in lowered:
            hits.append(term)
    return hits


def _build_lesion_sentence(region_name: str, long_axis_mm: float) -> str:
    size_phrase = "small" if 0 < long_axis_mm <= 20 else "focal"
    return f"A {size_phrase} hypoattenuating lesion is present in the {region_name}."


def _build_duct_sentence() -> str:
    return "Pancreatic duct dilation can be reviewed in relation to the lesion and surrounding pancreatic anatomy."


def _safe_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        f = float(value)
        if math.isnan(f):
            return None
        return f
    except Exception:
        return None


def _build_summary(metadata: Dict[str, Any], findings: List[Dict[str, Any]]) -> str:
    if metadata.get("narrative_report"):
        return str(metadata["narrative_report"])

    if findings:
        first = findings[0]["sentence"]
        return f"{first} Key images are linked below for rapid visual correlation."

    return "No structured lesion finding was generated for this case."


def _build_structured_impression(findings: List[Dict[str, Any]]) -> List[str]:
    if not findings:
        return ["No lesion-linked interactive findings were generated from the available labels."]
    return [f["sentence"] for f in findings]


def generate_case_report_json(case_id: str) -> Dict[str, Any]:
    paths = get_case_paths(case_id)
    ct_path = paths["ct_path"]
    combined_labels_path = paths["combined_labels_path"]

    if not os.path.exists(ct_path):
        raise FileNotFoundError(f"CT file not found: {ct_path}")
    if not os.path.exists(combined_labels_path):
        raise FileNotFoundError(f"Combined label map not found: {combined_labels_path}")

    metadata = load_case_metadata(case_id)
    ct_arr, label_arr, spacing_zyx = _read_ras_arrays(ct_path, combined_labels_path)

    findings: List[Dict[str, Any]] = []

    pancreatic_lesion_mask = _extract_binary_mask(label_arr, "pancreatic_lesion")
    if pancreatic_lesion_mask.sum() > 0:
        region_name = _infer_pancreatic_region(label_arr, pancreatic_lesion_mask)
        long_axis_mm = _approx_long_axis_mm(pancreatic_lesion_mask, spacing_zyx)
        sentence = _build_lesion_sentence(region_name, long_axis_mm)
        key_images = _render_overlay_and_zoom(ct_arr, pancreatic_lesion_mask, color="red")
        centroid = _centroid_zyx(pancreatic_lesion_mask)

        findings.append({
            "id": "pancreatic-lesion-1",
            "organ": "pancreas",
            "title": "Pancreatic lesion",
            "sentence": sentence,
            "kind": "lesion",
            "region_name": region_name,
            "long_axis_mm": long_axis_mm,
            "slice_index": key_images["slice_index"],
            "bbox": key_images["bbox"],
            "centroid_zyx": centroid,
            "terms": _term_hits(sentence),
            "images": {
                "overlay": key_images["overlay_image"],
                "zoom": key_images["zoom_image"],
            },
            "viewerLink": {
                "mode": "2D",
                "sliceIndex": key_images["slice_index"],
                "centroidZYX": centroid,
            },
        })

    pancreatic_duct_mask = _extract_binary_mask(label_arr, "pancreatic_duct")
    if pancreatic_duct_mask.sum() > 0:
        sentence = _build_duct_sentence()
        key_images = _render_overlay_and_zoom(ct_arr, pancreatic_duct_mask, color="gold")
        centroid = _centroid_zyx(pancreatic_duct_mask)

        findings.append({
            "id": "pancreatic-duct-1",
            "organ": "pancreas",
            "title": "Pancreatic duct review",
            "sentence": sentence,
            "kind": "duct",
            "region_name": "pancreatic duct",
            "long_axis_mm": None,
            "slice_index": key_images["slice_index"],
            "bbox": key_images["bbox"],
            "centroid_zyx": centroid,
            "terms": _term_hits(sentence),
            "images": {
                "overlay": key_images["overlay_image"],
                "zoom": key_images["zoom_image"],
            },
            "viewerLink": {
                "mode": "2D",
                "sliceIndex": key_images["slice_index"],
                "centroidZYX": centroid,
            },
        })

    glossary_payload = {
        term: {
            "label": val["label"],
            "definition": val["definition"],
            "example": val["example"],
        }
        for term, val in GLOSSARY.items()
    }

    age = _safe_number(metadata.get("age"))
    report = {
        "caseId": str(case_id),
        "pantsId": paths["pants_id"],
        "patient": {
            "age": age,
            "sex": metadata.get("sex"),
        },
        "study": {
            "scannerModel": metadata.get("scanner_model"),
            "studyDate": metadata.get("study_date"),
        },
        "report": {
            "summary": _build_summary(metadata, findings),
            "structuredImpression": _build_structured_impression(findings),
            "narrativeReport": metadata.get("narrative_report") or "",
            "structuredReport": metadata.get("structured_report") or "",
        },
        "findings": findings,
        "glossary": glossary_payload,
        "availableViews": ["2D", "3D"],
        "pdfExportUrl": f"/api/get-report/{case_id}",
    }

    return report